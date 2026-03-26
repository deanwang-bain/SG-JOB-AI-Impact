#!/usr/bin/env python3
"""
parse_wages.py — Extract and match wage data to SSOC occupations.

Reads MOM Occupational Wage Survey Excel files (2024), extracts median wages,
and fuzzy-matches occupation titles to SSOC 2024 codes.

The MOM files use SSOC 2020 codes but we match by occupation title to SSOC 2024.

Outputs:
- wages.csv: median monthly/annual wages per occupation with match confidence
"""

import json
import csv
from pathlib import Path
import openpyxl
from rapidfuzz import fuzz, process

RAW_WAGES_DIR = Path("raw/mom_wages")
OCCUPATIONS_JSON = Path("occupations.json")
OUTPUT_CSV = Path("wages.csv")


def load_occupations() -> dict:
    """Load SSOC occupations. Returns dict: {ssoc_code: {title, ...}}"""
    if not OCCUPATIONS_JSON.exists():
        print(f"✗ Error: {OCCUPATIONS_JSON} not found")
        print("  Run: uv run python parse_ssoc.py")
        return {}
    
    with open(OCCUPATIONS_JSON) as f:
        occupations = json.load(f)
    
    return {occ['ssoc_code']: occ for occ in occupations}


def extract_wages_from_excel(file_path: Path) -> list[dict]:
    """
    Extract wage data from a single MOM OWS Excel file (2024 format).
    
    Expected format:
    - Column A: Row number
    - Column B: SSOC 2020 code (5-digit)
    - Column C: Occupation title
    - Column D: Basic Wage ($)
    - Column E: Gross Wage ($)
    
    Returns list of dicts with keys:
    - ssoc_2020_code: original SSOC 2020 code from file
    - occupation_title: extracted occupation name
    - basic_wage: median monthly basic wage
    - gross_wage: median monthly gross wage
    """
    wages = []
    
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        
        for sheet_name in wb.sheetnames:
            # Skip Contents and metadata sheets
            if 'Content' in sheet_name or sheet_name in ['Publication Criteria', 'List A', 'List B']:
                continue
            
            sheet = wb[sheet_name]
            
            # Look for data rows
            # Data typically starts after row 8-10, with structure:
            # Col A: row number, Col B: SSOC code, Col C: Occupation, Col D: Basic wage, Col E: Gross wage
            
            for row_idx in range(1, sheet.max_row + 1):
                # Get row values
                row_num_cell = sheet.cell(row=row_idx, column=1)  # Column A
                ssoc_cell = sheet.cell(row=row_idx, column=2)     # Column B
                occupation_cell = sheet.cell(row=row_idx, column=3)  # Column C
                basic_wage_cell = sheet.cell(row=row_idx, column=4)  # Column D
                gross_wage_cell = sheet.cell(row=row_idx, column=5)  # Column E
                
                ssoc_code = ssoc_cell.value
                occupation = occupation_cell.value
                basic_wage = basic_wage_cell.value
                gross_wage = gross_wage_cell.value
                
                # Skip if no SSOC code or occupation
                if not ssoc_code or not occupation:
                    continue
                
                # Skip if SSOC code is not numeric (header rows)
                if not isinstance(ssoc_code, (int, float)):
                    # Try to parse as string number
                    try:
                        ssoc_code = int(str(ssoc_code))
                    except (ValueError, TypeError):
                        continue
                
                # Skip if occupation is not a string
                if not isinstance(occupation, str):
                    continue
                
                # Clean occupation title
                occupation = occupation.strip()
                
                # Skip category headers (ALL CAPS without specific details)
                if occupation.isupper() and len(occupation.split()) <= 2:
                    continue
                
                # Skip if no wage data
                if not basic_wage and not gross_wage:
                    continue
                
                # Handle 's' (suppressed) values
                if basic_wage == 's':
                    basic_wage = None
                if gross_wage == 's':
                    gross_wage = None
                
                # Convert wages to numbers
                try:
                    basic_wage = float(basic_wage) if basic_wage else None
                    gross_wage = float(gross_wage) if gross_wage else None
                except (ValueError, TypeError):
                    continue
                
                # Only add if we have at least one valid wage
                if basic_wage or gross_wage:
                    wages.append({
                        'ssoc_2020_code': int(ssoc_code),
                        'occupation_title': occupation,
                        'basic_wage': basic_wage,
                        'gross_wage': gross_wage,
                    })
        
        wb.close()
    
    except Exception as e:
        print(f"  ⚠ Error reading {file_path.name}: {e}")
    
    return wages


def fuzzy_match_to_ssoc(wages: list[dict], occupations: dict) -> list[dict]:
    """
    Fuzzy-match wage data to SSOC 2024 occupations by title.
    
    Returns list of dicts with matched SSOC codes and confidence scores.
    """
    # Build lookup: list of (title, ssoc_code) for matching
    ssoc_titles = [(occ['title'], code) for code, occ in occupations.items()]
    
    matched = []
    
    for wage in wages:
        title = wage['occupation_title']
        
        # Find best match using fuzzy matching
        result = process.extractOne(
            title,
            [t[0] for t in ssoc_titles],
            scorer=fuzz.token_sort_ratio
        )
        
        if result:
            matched_title, score, idx = result
            ssoc_code = ssoc_titles[idx][1]
            
            # Prefer basic wage, fall back to gross wage
            monthly_wage = wage['basic_wage'] if wage['basic_wage'] else wage['gross_wage']
            
            if monthly_wage:
                matched.append({
                    'ssoc_code': ssoc_code,
                    'occupation_title': wage['occupation_title'],
                    'matched_ssoc_title': matched_title,
                    'ssoc_2020_code': wage['ssoc_2020_code'],
                    'median_monthly_wage': monthly_wage,
                    'median_annual_wage': monthly_wage * 12,
                    'match_confidence': score / 100.0,  # Normalize to 0-1
                })
    
    return matched


def save_wages(wages: list[dict], min_confidence: float = 0.60):
    """Save wages to CSV, filtering low-confidence matches."""
    # Filter by minimum confidence
    filtered = [w for w in wages if w['match_confidence'] >= min_confidence]
    
    print(f"\nFiltering matches below {min_confidence:.0%} confidence...")
    print(f"  Kept: {len(filtered)}/{len(wages)} matches")
    
    if not filtered:
        print("⚠ No high-confidence matches found. Lowering threshold...")
        min_confidence = 0.50
        filtered = [w for w in wages if w['match_confidence'] >= min_confidence]
        print(f"  With {min_confidence:.0%} threshold: {len(filtered)} matches")
    
    # Sort by SSOC code
    filtered.sort(key=lambda x: (x['ssoc_code'], -x['match_confidence']))
    
    # Remove duplicates (keep highest confidence match for each SSOC code)
    deduped = {}
    for w in filtered:
        code = w['ssoc_code']
        if code not in deduped or w['match_confidence'] > deduped[code]['match_confidence']:
            deduped[code] = w
    
    wages = list(deduped.values())
    wages.sort(key=lambda x: x['ssoc_code'])
    
    # Save CSV
    if wages:
        with open(OUTPUT_CSV, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=wages[0].keys())
            writer.writeheader()
            writer.writerows(wages)
        print(f"✓ Saved {len(wages)} wage records to {OUTPUT_CSV}")
    
    # Statistics
    if wages:
        avg_confidence = sum(w['match_confidence'] for w in wages) / len(wages)
        print(f"\nMatch statistics:")
        print(f"  Average confidence: {avg_confidence:.2%}")
        print(f"  High confidence (≥0.8): {sum(1 for w in wages if w['match_confidence'] >= 0.8)}")
        print(f"  Medium confidence (0.6-0.8): {sum(1 for w in wages if 0.6 <= w['match_confidence'] < 0.8)}")
        print(f"  Low confidence (<0.6): {sum(1 for w in wages if w['match_confidence'] < 0.6)}")
        
        # Wage statistics
        monthly_wages = [w['median_monthly_wage'] for w in wages]
        print(f"\nWage statistics:")
        print(f"  Median monthly: ${sorted(monthly_wages)[len(monthly_wages)//2]:,.0f}")
        print(f"  Min: ${min(monthly_wages):,.0f} | Max: ${max(monthly_wages):,.0f}")
    
    return wages


def main():
    print("MOM Wage Data Parser")
    print("=" * 60)
    
    # Load SSOC occupations
    print("Loading SSOC occupations...")
    occupations = load_occupations()
    if not occupations:
        return
    print(f"✓ Loaded {len(occupations)} SSOC occupations")
    
    # Check for wage files
    if not RAW_WAGES_DIR.exists():
        print(f"✗ Error: {RAW_WAGES_DIR} not found")
        print("  Run: uv run python fetch_data.py")
        return
    
    wage_files = list(RAW_WAGES_DIR.glob("*.xlsx")) + list(RAW_WAGES_DIR.glob("*.xls"))
    
    if not wage_files:
        print(f"✗ Error: No Excel files found in {RAW_WAGES_DIR}")
        print("  Run: uv run python fetch_data.py")
        return
    
    print(f"\nFound {len(wage_files)} wage files")
    
    # Extract wages from all files
    all_wages = []
    for i, file_path in enumerate(wage_files, 1):
        print(f"[{i}/{len(wage_files)}] Parsing {file_path.name}...")
        wages = extract_wages_from_excel(file_path)
        all_wages.extend(wages)
        print(f"  ✓ Extracted {len(wages)} wage entries")
    
    print(f"\n✓ Total wage entries extracted: {len(all_wages)}")
    
    if not all_wages:
        print("⚠ No wage data extracted. Manual intervention may be required.")
        return
    
    # Fuzzy match to SSOC
    print("\nMatching wages to SSOC occupations...")
    matched = fuzzy_match_to_ssoc(all_wages, occupations)
    
    # Save outputs
    final_wages = save_wages(matched)
    
    print("\n" + "=" * 60)
    print("✓ Wage parsing complete!")
    
    # Show examples
    if final_wages:
        # Sort by confidence to show best matches
        best_matches = sorted(final_wages, key=lambda x: -x['match_confidence'])[:5]
        
        print("\nExample high-confidence matches:")
        for i, w in enumerate(best_matches, 1):
            print(f"\n{i}. {w['matched_ssoc_title']} ({w['ssoc_code']})")
            print(f"   MOM title: {w['occupation_title']}")
            print(f"   Monthly: ${w['median_monthly_wage']:,.0f} | Annual: ${w['median_annual_wage']:,.0f}")
            print(f"   Match confidence: {w['match_confidence']:.1%}")


if __name__ == "__main__":
    main()
