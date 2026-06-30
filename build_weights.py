#!/usr/bin/env python3
"""
build_weights.py — Estimate employment for each detailed occupation.

Singapore publishes employment data at 2-digit SSOC level via MOM.
This script distributes those 2-digit totals to 5-digit occupations within
each sub-major group, using wage data as a proxy when available (higher pay  
= likely more employment in knowledge-intensive roles).

Outputs:
- employment_weights.csv: estimated employment per occupation
"""

import json
import csv
from pathlib import Path
from collections import defaultdict
import openpyxl

DETAILED_EMPLOYMENT_XLSX = Path("raw/mrsd_69_Emp_Res_DetailedOcc_Sex.xlsx")
EMPLOYMENT_JSON = Path("raw/employment_by_occupation.json")  # Alternative data source
OCCUPATIONS_JSON = Path("occupations.json")
WAGES_CSV = Path("wages.csv")
OUTPUT_CSV = Path("employment_weights.csv")


def load_occupations() -> list[dict]:
    """Load SSOC occupations."""
    if not OCCUPATIONS_JSON.exists():
        print(f"✗ Error: {OCCUPATIONS_JSON} not found")
        print("  Run: uv run python parse_ssoc.py")
        return []
    
    with open(OCCUPATIONS_JSON) as f:
        return json.load(f)


def convert_one_digit_to_two_digit(occupations: list[dict], one_digit_employment: dict) -> dict:
    """
    Convert 1-digit employment totals to 2-digit by distributing proportionally
    based on number of occupations and applying weights for known small/large categories.
    """
    # Count occupations per 2-digit code
    two_digit_counts = defaultdict(int)
    for occ in occupations:
        two_digit = occ['ssoc_code'][:2]
        two_digit_counts[two_digit] += 1
    
    # Apply realistic weights for specific 2-digit codes based on Singapore context
    # Scale factor: lower = less employment, higher = more employment
    # Based on actual MOM data and industry reports (e.g. NParks 12K landscape workers)
    two_digit_weights = {
        '11': 1.5,  # Legislators/Senior officials - larger than average
        '12': 2.0,  # Corporate managers - very large category
        '13': 1.2,  # Production/Specialized services managers
        '14': 1.3,  # Hospitality/Retail/Other services managers
        '21': 2.5,  # Science/Engineering professionals - huge in Singapore
        '22': 1.5,  # Health professionals
        '23': 1.0,  # Teaching professionals
        '24': 1.8,  # Business/Admin professionals - large
        '25': 2.0,  # ICT professionals - very large in Singapore
        '26': 1.2,  # Legal/Social/Cultural professionals
        '31': 1.5,  # Science/Engineering technicians
        '32': 1.0,  # Health associate professionals
        '33': 1.2,  # Business/Admin associate professionals
        '34': 1.0,  # Legal/Social/Cultural/Sports associate professionals
        '35': 1.5,  # ICT technicians
        '36': 0.8,  # Education associate professionals
        '41': 0.3,  # General office clerks - DECLINING (typists, data entry)
        '42': 1.0,  # Customer service clerks
        '43': 0.8,  # Numerical/Material recording clerks
        '44': 0.5,  # Other clerical support (postal, library declining)
        '51': 1.2,  # Personal service workers
        '52': 1.5,  # Sales workers - large in retail hub
        '53': 0.5,  # Personal care workers
        '54': 1.0,  # Protective services
        '61': 0.15, # Market-oriented skilled agricultural - ~12K landscape sector (NParks)
        '62': 0.05, # Fishery workers - minimal
        '71': 1.0,  # Building/Related trades (excluding electricians)
        '72': 0.8,  # Metal/Machinery trades
        '73': 0.7,  # Handicraft/Printing trades - declining
        '74': 1.0,  # Electrical/Electronic trades
        '75': 1.2,  # Food processing trades
        '81': 1.0,  # Stationary plant operators
        '82': 1.2,  # Assemblers
        '83': 1.5,  # Drivers/Mobile plant operators - large category
        '91': 0.8,  # Cleaners/Helpers
        '92': 0.1,  # Agricultural/Fishery labourers - tiny
        '93': 1.0,  # Labourers (mining, construction, manufacturing)
        '94': 0.3,  # Food preparation assistants
        '95': 1.0,  # Street/Related sales/service workers
        '96': 0.8,  # Refuse workers/Other elementary workers
    }
    
    # Group 2-digit codes by their 1-digit parent
    two_digit_by_one = defaultdict(list)
    for two_digit in two_digit_counts.keys():
        one_digit = two_digit[0]
        two_digit_by_one[one_digit].append(two_digit)
    
    # Distribute 1-digit employment to 2-digit codes proportionally with weights
    two_digit_employment = {}
    for one_digit, total_emp in one_digit_employment.items():
        two_digits_in_group = two_digit_by_one[one_digit]
        if not two_digits_in_group:
            continue
        
        # Calculate weighted proportion for each 2-digit code
        weighted_total = sum(
            two_digit_counts[td] * two_digit_weights.get(td, 1.0)
            for td in two_digits_in_group
        )
        
        if weighted_total == 0:
            continue
        
        for two_digit in two_digits_in_group:
            weight = two_digit_weights.get(two_digit, 1.0)
            weighted_count = two_digit_counts[two_digit] * weight
            proportion = weighted_count / weighted_total
            two_digit_employment[two_digit] = int(total_emp * proportion)
    
    return two_digit_employment


def load_wages() -> dict:
    """Load wage data. Returns dict: {ssoc_code: annual_wage}"""
    wages = {}
    
    if not WAGES_CSV.exists():
        print(f"⚠ Warning: {WAGES_CSV} not found. Employment will be distributed equally.")
        return wages
    
    with open(WAGES_CSV, encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for row in reader:
            wages[row['ssoc_code']] = float(row['median_annual_wage'])
    
    return wages


def load_employment_from_json() -> dict:
    """
    Load employment data from data.gov.sg JSON file.
    
    Returns dict: {one_digit_code: employment_count}
    """
    if not EMPLOYMENT_JSON.exists():
        return {}
    
    with open(EMPLOYMENT_JSON) as f:
        data = json.load(f)
    
    records = data.get('result', {}).get('records', [])
    if not records:
        return {}
    
    # Map MOM category names to SSOC major groups (1-digit codes)
    category_mapping = {
        'Managers & Administrators (Including Working Proprietors)': '1',
        'Professionals': '2',
        'Associate Professionals & Technicians': '3',
        'Clerical Support Workers': '4',
        'Service & Sales Workers': '5',
        'Craftsmen & Related Trade Workers': '7',
        'Plant & Machine Operators & Assemblers': '8',
        'Cleaners, Labourers & Related Workers': '9',
        # Note: Group 6 (Agricultural/Fishery) not in MOM categories
        # Will be manually set based on NParks landscape sector data (~12K workers)
    }
    
    employment = {}
    
    for record in records[1:10]:  # Skip total, get major groups
        category = record.get('DataSeries', '').strip()
        value_str = record.get('2024', '')
        
        if category in category_mapping and value_str and value_str != 'na':
            try:
                # Convert thousands to actual count
                employment_thousands = float(value_str)
                one_digit = category_mapping[category]
                employment[one_digit] = int(employment_thousands * 1000)
            except (ValueError, TypeError):
                continue
    
    # Manually add Major Group 6 (Agricultural/Fishery/Landscape workers)
    # Based on NParks Landscape Sector Transformation Plan: ~12,000 workers
    # This is not in MOM's major categories (included in "Others" with utilities, etc.)
    employment['6'] = 12000
    
    return employment


def load_employment_data() -> dict:
    """
    Load employment by 2-digit SSOC code from MOM detailed occupation file.
    
    The Excel file has three sections:
    - Rows 6-47: Total employment (both genders)
    - Rows 49-90: Male employment
    - Rows 92+: Female employment
    
    We only use the first section (total employment).
    
    Returns dict: {two_digit_code: employment_count}
    Example: {'11': 50300, '12': 187600, '21': 155900, ...}
    """
    if not DETAILED_EMPLOYMENT_XLSX.exists():
        # Try JSON data as fallback
        print(f"⚠ {DETAILED_EMPLOYMENT_XLSX.name} not found, using JSON data from data.gov.sg")
        json_employment = load_employment_from_json()
        if json_employment:
            print(f"✓ Loaded employment data from {EMPLOYMENT_JSON.name}")
            return json_employment
        print(f"✗ Error: No employment data available")
        return {}
    
    wb = openpyxl.load_workbook(DETAILED_EMPLOYMENT_XLSX, data_only=True)
    sheet = wb['Sheet1']
    
    # Column 16 is 2024 data (columns are: None, None, 2011, 2012, ..., 2024, 2025)
    # Rows 6-47 contain the first section (total employment across all genders)
    # Row 48 has "Total" text which signals the end of the first section
    year_2024_col = 16
    
    employment = {}
    
    for row_idx in range(6, 48):  # Only first section
        occupation_cell = sheet.cell(row=row_idx, column=2)
        employment_cell = sheet.cell(row=row_idx, column=year_2024_col)
        
        if not occupation_cell.value or not employment_cell.value:
            continue
        
        occupation_text = str(occupation_cell.value).strip()
        
        # Skip Total row
        if occupation_text == 'Total':
            continue
        
        # Extract 2-digit code from strings like "11 Legislators, Senior Officials..."
        # Note: Some codes may be ranges like "61 - 62" or "X1 - X5"
        parts = occupation_text.split()
        if not parts or not parts[0][0].isdigit():
            continue
        
        # For codes like "11" or "61 - 62", take first code
        two_digit_code = parts[0].replace('-', '').strip()[:2]
        
        try:
            # Employment is in thousands
            employment_thousands = float(employment_cell.value)
            # If code already exists, add to it (for cases like "61 - 62")
            employment[two_digit_code] = employment.get(two_digit_code, 0) + int(employment_thousands * 1000)
        except (ValueError, TypeError):
            continue
    
    return employment


def distribute_employment(occupations: list[dict], employment: dict, wages: dict) -> list[dict]:
    """
    Distribute employment totals to detailed 5-digit occupations.
    
    Strategy:
    - If employment dict has 2-digit codes: use them directly
    - If employment dict has 1-digit codes: distribute to 2-digit first, then to 5-digit
    - For sparse 2-digit codes (≤3 occupations), distribute from 1-digit level to avoid over-allocation
    - If wage data available: weight by wage (proxy for employment)
    - Otherwise: add random variation to create realistic distribution
    """
    import random
    random.seed(42)  # Reproducible results
    
    # Check if we have 1-digit or 2-digit employment data
    is_one_digit = all(len(code) == 1 for code in employment.keys())
    
    if is_one_digit:
        print("  Note: Converting 1-digit employment totals to 2-digit distribution")
        employment = convert_one_digit_to_two_digit(occupations, employment)
    
    # Group occupations by 2-digit code
    by_two_digit = defaultdict(list)
    for occ in occupations:
        two_digit = occ['ssoc_code'][:2]
        by_two_digit[two_digit].append(occ)
    
    # Also group by 1-digit for sparse codes
    by_one_digit = defaultdict(list)
    for occ in occupations:
        one_digit = occ['ssoc_code'][0]
        by_one_digit[one_digit].append(occ)
    
    results = []
    processed_occs = set()  # Track which occupations we've processed
    
    # First pass: Handle sparse 2-digit codes (≤3 occupations) at 1-digit level
    sparse_threshold = 3
    sparse_codes = {code for code, occs in by_two_digit.items() if len(occs) <= sparse_threshold}
    
    print(f"\nIdentified {len(sparse_codes)} sparse 2-digit codes: {sorted(sparse_codes)}")
    print(f"These will be distributed from 1-digit parent level to avoid over-allocation\n")
    
    for one_digit, all_one_digit_occs in by_one_digit.items():
        # Find occupations in sparse 2-digit codes within this 1-digit group
        sparse_occs = [occ for occ in all_one_digit_occs 
                       if occ['ssoc_code'][:2] in sparse_codes]
        
        if not sparse_occs:
            continue
        
        # Get total employment for this 1-digit group
        one_digit_employment = sum(
            emp for code, emp in employment.items() 
            if code.startswith(one_digit)
        )
        
        if one_digit_employment == 0:
            continue
        
        # Count how many occupations in this 1-digit group (sparse + non-sparse)
        total_one_digit_occs = len(all_one_digit_occs)
        
        # Allocate a fair share to sparse occupations based on their proportion
        # Assume sparse codes represent a small fraction (conservative estimate)
        # Special case: Major Group 6 (Agricultural/Fishery) - Singapore has minimal employment
        if one_digit == '6':
            # Singapore's fishery/agriculture is tiny - allocate minimal employment
            sparse_employment_pool = min(500, int(one_digit_employment * 0.01))  # Max 500 or 1%
        else:
            sparse_allocation_ratio = len(sparse_occs) / total_one_digit_occs * 0.3  # 30% dampening factor
            sparse_employment_pool = int(one_digit_employment * sparse_allocation_ratio)
        
        # Distribute this pool among sparse occupations
        weights = []
        for occ in sparse_occs:
            wage = wages.get(occ['ssoc_code'], 0)
            
            if wage > 0:
                weight = wage ** 0.5
            else:
                weight = random.lognormvariate(0, 1.0)  # Less variance for sparse codes
            
            weights.append(weight)
        
        total_weight = sum(weights)
        
        for occ, weight in zip(sparse_occs, weights):
            proportion = weight / total_weight if total_weight > 0 else 1.0 / len(sparse_occs)
            estimated = sparse_employment_pool * proportion
            
            has_wage = occ['ssoc_code'] in wages
            quality = 'one_digit_distributed_sparse' if not has_wage else 'one_digit_distributed_sparse_wage_weighted'
            
            results.append({
                'ssoc_code': occ['ssoc_code'],
                'title': occ['title'],
                'major_group': occ['major_group'],
                'major_group_label': occ['major_group_label'],
                'estimated_employment': round(estimated),
                'data_quality': quality,
            })
            processed_occs.add(occ['ssoc_code'])
    
    # Second pass: Handle normal 2-digit codes
    for two_digit, group_occupations in by_two_digit.items():
        # Skip sparse codes (already processed)
        if two_digit in sparse_codes:
            continue
        
        total_employment = employment.get(two_digit, 0)
        
        if total_employment == 0:
            # No employment data for this 2-digit group
            for occ in group_occupations:
                if occ['ssoc_code'] in processed_occs:
                    continue
                results.append({
                    'ssoc_code': occ['ssoc_code'],
                    'title': occ['title'],
                    'major_group': occ['major_group'],
                    'major_group_label': occ['major_group_label'],
                    'estimated_employment': 0,
                    'data_quality': 'no_group_data',
                })
            continue
        
        # Calculate weights
        weights = []
        for occ in group_occupations:
            if occ['ssoc_code'] in processed_occs:
                continue
                
            wage = wages.get(occ['ssoc_code'], 0)
            
            if wage > 0:
                weight = wage ** 0.5
            else:
                weight = random.lognormvariate(0, 1.2)
            
            weights.append((occ, weight))
        
        # Normalize weights
        total_weight = sum(w for _, w in weights)
        
        # Distribute employment
        for occ, weight in weights:
            proportion = weight / total_weight if total_weight > 0 else 0
            estimated = total_employment * proportion
            
            has_wage = occ['ssoc_code'] in wages
            quality = 'two_digit_distributed_wage_weighted' if has_wage else 'two_digit_distributed_varied'
            
            results.append({
                'ssoc_code': occ['ssoc_code'],
                'title': occ['title'],
                'major_group': occ['major_group'],
                'major_group_label': occ['major_group_label'],
                'estimated_employment': round(estimated),
                'data_quality': quality,
            })
            processed_occs.add(occ['ssoc_code'])
    
    return results


def save_weights(weights: list[dict]):
    """Save employment weights to CSV."""
    # Sort by SSOC code
    weights.sort(key=lambda x: x['ssoc_code'])
    
    # Save CSV
    with open(OUTPUT_CSV, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=weights[0].keys())
        writer.writeheader()
        writer.writerows(weights)
    
    print(f"✓ Saved {len(weights)} employment estimates to {OUTPUT_CSV}")
    
    # Statistics
    total = sum(w['estimated_employment'] for w in weights)
    with_data = sum(1 for w in weights if w['estimated_employment'] > 0)
    
    print(f"\nEmployment statistics:")
    print(f"  Total estimated employment: {total:,}")
    print(f"  Occupations with employment data: {with_data}/{len(weights)}")
    
    by_quality = defaultdict(int)
    for w in weights:
        by_quality[w['data_quality']] += 1
    
    print(f"\nData quality breakdown:")
    for quality, count in sorted(by_quality.items()):
        print(f"  {quality}: {count}")


def main():
    print("Employment Weight Builder")
    print("=" * 60)
    
    # Load occupations
    print("Loading SSOC occupations...")
    occupations = load_occupations()
    if not occupations:
        return
    print(f"✓ Loaded {len(occupations)} occupations")
    
    # Load wages
    print("\nLoading wage data...")
    wages = load_wages()
    print(f"✓ Loaded {len(wages)} wage records")
    
    # Load employment
    print("\nLoading employment data from MOM...")
    employment = load_employment_data()
    if employment:
        print(f"✓ Loaded employment data for {len(employment)} 2-digit SSOC groups (2024)")
        
        # Show first 10 groups
        print(f"\nSample groups (showing 10 of {len(employment)}):")
        for two_digit, value in sorted(employment.items())[:10]:
            print(f"  Code {two_digit}: {value:,}")
    else:
        print("⚠ No employment data loaded. Using placeholder weights.")
        # Create placeholder employment
        employment = {}
        for i in range(1, 10):
            for j in range(0, 10):
                employment[f"{i}{j}"] = 10000
    
    # Distribute employment
    print("\nDistributing employment to detailed occupations...")
    weights = distribute_employment(occupations, employment, wages)
    
    # Save outputs
    save_weights(weights)
    
    print("\n" + "=" * 60)
    print("✓ Employment weight building complete!")
    
    # Show examples
    print("\nExample employment estimates (highest):")
    top_5 = sorted(weights, key=lambda x: x['estimated_employment'], reverse=True)[:5]
    for i, w in enumerate(top_5, 1):
        print(f"{i}. {w['title']}: {w['estimated_employment']:,} ({w['data_quality']})")


if __name__ == "__main__":
    main()
