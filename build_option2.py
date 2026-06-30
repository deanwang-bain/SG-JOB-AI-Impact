#!/usr/bin/env python3
"""
Build Option 2: 441 detailed occupations distributed from REAL 2-digit MOM data.

Key improvements over old approach:
- Start from REAL 2-digit MOM employment (not 1-digit estimates)
- Wage-weighted distribution within each 2-digit category
- Calibrate to exactly match MOM total (2,346,000)
- Clear data quality labels
"""

import json
import csv
import openpyxl
from pathlib import Path
from collections import defaultdict

# Paths
OCCUPATIONS_JSON = Path("occupations.json")
WAGES_CSV = Path("wages.csv")
SCORES_JSON = Path("scores.json")
MOM_EXCEL = Path("raw/mrsd_69_Emp_Res_DetailedOcc_Sex.xlsx")
OUTPUT_JSON = Path("docs/data.json")

def load_mom_2digit_employment():
    """Load REAL 2-digit employment from MOM Excel."""
    print("Loading REAL 2-digit employment from MOM Excel...")
    
    wb = openpyxl.load_workbook(MOM_EXCEL, data_only=True)
    sheet = wb['Sheet1']
    
    # Find 2024 column
    col_2024 = None
    for col in range(1, 20):
        if sheet.cell(row=4, column=col).value == 2024:
            col_2024 = col
            break
    
    employment = {}
    for row in range(6, 50):
        occupation_cell = sheet.cell(row=row, column=2)
        employment_cell = sheet.cell(row=row, column=col_2024)
        
        if not occupation_cell.value:
            continue
        
        occ_text = str(occupation_cell.value).strip()
        
        if len(occ_text) >= 2 and occ_text[:2].isdigit():
            code = occ_text[:2]
            
            # Handle "61 - 62" merged category
            if ' - ' in code or ' - ' in occ_text[:5]:
                code = '61'
            
            try:
                employment_thousands = float(employment_cell.value)
                employment_count = int(employment_thousands * 1000)
                
                if code in employment:
                    employment[code] += employment_count
                else:
                    employment[code] = employment_count
            except (ValueError, TypeError):
                continue
    
    total = sum(employment.values())
    print(f"✓ Loaded {len(employment)} 2-digit categories")
    print(f"  Total employment: {total:,}")
    return employment


def load_occupations_and_wages():
    """Load 5-digit occupations and wages."""
    print("Loading 5-digit occupations and wages...")
    
    with open(OCCUPATIONS_JSON) as f:
        occupations = json.load(f)
    
    wages = {}
    try:
        with open(WAGES_CSV) as f:
            reader = csv.DictReader(f)
            for row in reader:
                try:
                    wages[row['ssoc_code']] = float(row['median_annual_wage'])
                except (ValueError, KeyError):
                    pass
    except FileNotFoundError:
        print("⚠ Warning: wages.csv not found")
    
    print(f"✓ Loaded {len(occupations)} 5-digit occupations")
    print(f"✓ Loaded {len(wages)} wage records")
    
    return occupations, wages


def load_ai_scores():
    """Load AI exposure scores."""
    print("Loading AI exposure scores...")
    
    if not SCORES_JSON.exists():
        print("⚠ Warning: scores.json not found")
        return {}
    
    with open(SCORES_JSON) as f:
        scores_list = json.load(f)
        scores = {s['ssoc_code']: s for s in scores_list}
    
    print(f"✓ Loaded {len(scores)} AI exposure scores")
    return scores


def distribute_2digit_to_5digit(occupations, employment_2digit, wages):
    """
    Distribute REAL 2-digit employment to 5-digit occupations.
    
    Strategy:
    1. Group 5-digit occupations by 2-digit parent
    2. For occupations with wage data: distribute proportionally by wage
    3. For occupations without wage data: equal distribution of remaining pool
    4. Calibrate final totals to match exactly
    """
    print("Distributing REAL 2-digit employment to 5-digit occupations...")
    
    # Group by 2-digit
    by_two_digit = defaultdict(list)
    for occ in occupations:
        two_digit = occ['ssoc_code'][:2]
        by_two_digit[two_digit].append(occ)
    
    results = []
    uncalibrated_total = 0
    
    for two_digit, total_emp in employment_2digit.items():
        occs = by_two_digit.get(two_digit, [])
        if not occs:
            continue
        
        # Separate occupations with and without wage data
        with_wages = [(occ, wages.get(occ['ssoc_code'], 0)) for occ in occs if wages.get(occ['ssoc_code'], 0) > 0]
        without_wages = [occ for occ in occs if wages.get(occ['ssoc_code'], 0) == 0]
        
        if with_wages:
            # Use 80% of total for wage-weighted distribution
            # Use 20% for equal distribution among no-wage occupations
            wage_pool = int(total_emp * 0.8)
            equal_pool = int(total_emp * 0.2)
            
            # Wage-weighted distribution
            total_wage = sum(w for _, w in with_wages)
            for occ, wage in with_wages:
                proportion = wage / total_wage
                estimated = int(wage_pool * proportion)
                results.append({
                    'occ': occ,
                    'employment': estimated,
                    'method': 'wage_weighted_real_2digit'
                })
                uncalibrated_total += estimated
            
            # Equal distribution for no-wage
            if without_wages:
                per_occ = equal_pool // len(without_wages)
                for occ in without_wages:
                    results.append({
                        'occ': occ,
                        'employment': per_occ,
                        'method': 'equal_share_real_2digit'
                    })
                    uncalibrated_total += per_occ
        else:
            # All equal if no wage data
            per_occ = total_emp // len(occs)
            for occ in occs:
                results.append({
                    'occ': occ,
                    'employment': per_occ,
                    'method': 'equal_real_2digit'
                })
                uncalibrated_total += per_occ
    
    print(f"✓ Distributed to {len(results)} occupations")
    print(f"  Uncalibrated total: {uncalibrated_total:,}")
    
    return results, uncalibrated_total


def calibrate_employment(results, target_total):
    """Calibrate employment to match exact MOM total."""
    print(f"Calibrating to match MOM total: {target_total:,}...")
    
    current_total = sum(r['employment'] for r in results)
    calibration_factor = target_total / current_total
    
    calibrated_results = []
    calibrated_total = 0
    
    for r in results:
        calibrated_emp = int(r['employment'] * calibration_factor)
        calibrated_results.append({
            'occ': r['occ'],
            'employment': calibrated_emp,
            'method': r['method']
        })
        calibrated_total += calibrated_emp
    
    # Adjust largest occupation to exactly match target
    if calibrated_total != target_total:
        diff = target_total - calibrated_total
        largest = max(calibrated_results, key=lambda x: x['employment'])
        largest['employment'] += diff
        calibrated_total = target_total
    
    print(f"✓ Calibrated to: {calibrated_total:,} (100% match)")
    return calibrated_results


def build_dataset(results, scores):
    """Build final dataset with all metadata."""
    print("Building final dataset...")
    
    data = []
    for r in sorted(results, key=lambda x: x['occ']['ssoc_code']):
        occ = r['occ']
        score_data = scores.get(occ['ssoc_code'], {})
        
        entry = {
            'ssoc_code': occ['ssoc_code'],
            'title': occ['title'],
            'slug': occ['slug'],
            'major_group': occ['major_group'],
            'category': occ['major_group_label'].lower().replace(' ', '-').replace(',', ''),
            'category_label': occ['major_group_label'],
            'jobs': r['employment'],
            'pay': None,  # Will be added if available
            'education': None,
            'exposure': score_data.get('exposure') or score_data.get('exposure_score'),
            'exposure_rationale': score_data.get('rationale') or score_data.get('exposure_rationale'),
            'ssoc_url': "https://go.gov.sg/ssoc-search-engine",
            'data_quality': r['method'],  # Clear labeling of estimation method
        }
        
        data.append(entry)
    
    # Add wage data
    try:
        with open(WAGES_CSV) as f:
            reader = csv.DictReader(f)
            wage_dict = {}
            for row in reader:
                try:
                    wage_dict[row['ssoc_code']] = float(row['median_annual_wage'])
                except (ValueError, KeyError):
                    pass
        
        for entry in data:
            if entry['ssoc_code'] in wage_dict:
                entry['pay'] = wage_dict[entry['ssoc_code']]
    except FileNotFoundError:
        pass
    
    print(f"✓ Built dataset with {len(data)} occupations")
    return data


def calculate_statistics(data):
    """Calculate summary statistics."""
    stats = {
        'total_occupations': len(data),
        'total_workforce': sum(d['jobs'] for d in data),
        'with_pay': sum(1 for d in data if d['pay']),
        'with_exposure': sum(1 for d in data if d['exposure'] is not None),
    }
    
    # Average exposure (unweighted)
    exposures = [d['exposure'] for d in data if d['exposure'] is not None]
    if exposures:
        stats['avg_exposure'] = sum(exposures) / len(exposures)
    
    # Weighted average exposure
    weighted_sum = 0
    total_weight = 0
    for d in data:
        if d['exposure'] is not None and d['jobs']:
            weighted_sum += d['exposure'] * d['jobs']
            total_weight += d['jobs']
    
    if total_weight > 0:
        stats['weighted_avg_exposure'] = weighted_sum / total_weight
    
    # PME stats
    pme_codes = ['1', '2', '3']
    pme_workforce = sum(d['jobs'] for d in data if d['major_group'] in pme_codes)
    if stats['total_workforce'] > 0:
        stats['pme_workforce'] = pme_workforce
        stats['pme_share'] = pme_workforce / stats['total_workforce']
    
    pme_exposures = [(d['exposure'], d['jobs']) for d in data 
                     if d['major_group'] in pme_codes and d['exposure'] is not None and d['jobs']]
    if pme_exposures:
        pme_weighted_sum = sum(exp * jobs for exp, jobs in pme_exposures)
        pme_total_weight = sum(jobs for _, jobs in pme_exposures)
        stats['pme_avg_exposure'] = pme_weighted_sum / pme_total_weight if pme_total_weight > 0 else None
    
    return stats


def main():
    print("=" * 70)
    print("Building Option 2: 441 Occupations with Real 2-digit Base")
    print("=" * 70)
    print()
    
    # Load data
    employment_2digit = load_mom_2digit_employment()
    occupations, wages = load_occupations_and_wages()
    scores = load_ai_scores()
    
    # Distribute
    results, uncalibrated_total = distribute_2digit_to_5digit(occupations, employment_2digit, wages)
    
    # Calibrate to exact MOM total
    target_total = sum(employment_2digit.values())
    results = calibrate_employment(results, target_total)
    
    # Build dataset
    data = build_dataset(results, scores)
    
    # Calculate statistics
    stats = calculate_statistics(data)
    
    # Create output
    output = {
        'metadata': {
            'generated': '2026-06-30',
            'version': '2.0',
            'granularity': '5-digit SSOC (distributed from real 2-digit)',
            'data_quality': 'Real 2-digit base, wage-weighted 5-digit distribution',
            'sources': [
                'MOM Detailed Occupation Employment 2024 (mrsd_69) - 2-digit REAL data',
                'MOM Occupational Wage Survey 2024 - for distribution weights',
                'SSOC 2020 Classification (SingStat)',
                'OpenAI GPT-4o AI Exposure Scoring'
            ]
        },
        'statistics': stats,
        'occupations': data
    }
    
    # Save
    with open(OUTPUT_JSON, 'w') as f:
        json.dump(output, f, indent=2)
    
    print(f"\n✓ Saved {len(data)} occupations to {OUTPUT_JSON}")
    
    # Summary
    print("\n" + "=" * 70)
    print("Dataset summary:")
    print(f"  Total occupations: {stats['total_occupations']}")
    print(f"  Total workforce: {stats['total_workforce']:,}")
    print(f"  With wage data: {stats['with_pay']} ({stats['with_pay']/stats['total_occupations']*100:.1f}%)")
    print(f"  With AI scores: {stats['with_exposure']} ({stats['with_exposure']/stats['total_occupations']*100:.1f}%)")
    print(f"  Average AI exposure: {stats.get('avg_exposure', 0):.2f}/10")
    print(f"  Weighted AI exposure: {stats.get('weighted_avg_exposure', 0):.2f}/10")
    
    # Show distribution methods
    by_method = defaultdict(int)
    for d in data:
        by_method[d['data_quality']] += 1
    
    print("\n  Distribution methods:")
    for method, count in sorted(by_method.items()):
        print(f"    {method}: {count} ({count/len(data)*100:.1f}%)")
    
    print("\n✓ Ready for visualization!")
    print("  Employment starts from REAL 2-digit MOM data")
    print("  5-digit distribution uses wage-weighting")


if __name__ == '__main__':
    main()
